import json
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from .models import Campaign, Lead
from .forms import CampaignForm
from .scraper import run_campaign_scrape


# ── SPA shell ────────────────────────────────────────────────────────────────
def index(request):
    import os
    from django.conf import settings
    dist = os.path.join(settings.BASE_DIR, 'leads', 'static', 'leads', 'dist', 'assets')
    css_files, js_files = [], []
    if os.path.exists(dist):
        for f in os.listdir(dist):
            if f.endswith('.css'):
                css_files.append(f'leads/dist/assets/{f}')
            elif f.endswith('.js'):
                js_files.append(f'leads/dist/assets/{f}')
    return render(request, 'leads/index.html', {'css_files': css_files, 'js_files': js_files})


# ── API helpers ───────────────────────────────────────────────────────────────
def campaign_to_dict(c):
    leads = list(c.leads.order_by('-ai_score').values(
        'id', 'name', 'title', 'company', 'email', 'linkedin_url', 'ai_score', 'status'
    ))
    return {
        'id': c.id,
        'name': c.name,
        'saas_description': c.saas_description,
        'target_audience': c.target_audience,
        'ai_search_queries': c.ai_search_queries,
        'created_at': c.created_at.isoformat(),
        'leads': leads,
        'lead_count': len(leads),
    }


# ── Campaign API ──────────────────────────────────────────────────────────────
@csrf_exempt
@require_http_methods(['GET', 'POST'])
def api_campaigns(request):
    if request.method == 'GET':
        campaigns = Campaign.objects.prefetch_related('leads').order_by('-created_at')
        return JsonResponse([campaign_to_dict(c) for c in campaigns], safe=False)

    data = json.loads(request.body)
    campaign = Campaign.objects.create(
        name=data['name'],
        saas_description=data['saas_description'],
        target_audience=data['target_audience'],
    )
    return JsonResponse(campaign_to_dict(campaign), status=201)


@require_http_methods(['GET'])
def api_campaign_detail(request, pk):
    campaign = get_object_or_404(Campaign, pk=pk)
    return JsonResponse(campaign_to_dict(campaign))


@csrf_exempt
@require_http_methods(['POST'])
def api_run_scrape(request, pk):
    campaign = get_object_or_404(Campaign, pk=pk)
    try:
        scraped = run_campaign_scrape(campaign)
        created = 0
        updated = 0
        for lead_data in scraped:
            if not lead_data.get('linkedin_url'):
                continue
            lead, is_new = Lead.objects.get_or_create(
                campaign=campaign,
                linkedin_url=lead_data['linkedin_url'],
                defaults={
                    'name': lead_data.get('name', ''),
                    'email': lead_data.get('email', ''),
                    'title': lead_data.get('title', ''),
                    'company': lead_data.get('company', ''),
                    'ai_score': lead_data.get('ai_score', 50),
                }
            )
            if not is_new:
                # always update score + email on re-scrape
                changed = False
                if lead_data.get('email') and not lead.email:
                    lead.email = lead_data['email']
                    changed = True
                if lead_data.get('ai_score') and lead_data['ai_score'] != lead.ai_score:
                    lead.ai_score = lead_data['ai_score']
                    changed = True
                if changed:
                    lead.save()
                    updated += 1
            else:
                created += 1
        return JsonResponse({'created': created, 'updated': updated, 'campaign': campaign_to_dict(campaign)})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(['PATCH'])
def api_update_lead(request, pk):
    lead = get_object_or_404(Lead, pk=pk)
    data = json.loads(request.body)
    lead.status = data.get('status', lead.status)
    lead.save()
    return JsonResponse({'id': lead.id, 'status': lead.status})


@require_http_methods(['GET'])
def api_export_excel(request, pk):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    campaign = get_object_or_404(Campaign, pk=pk)
    leads = campaign.leads.order_by('-ai_score')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Leads'

    headers = ['Name', 'Title', 'Company', 'Email', 'LinkedIn URL', 'AI Score', 'Status']
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row, lead in enumerate(leads, 2):
        ws.cell(row=row, column=1, value=lead.name)
        ws.cell(row=row, column=2, value=lead.title)
        ws.cell(row=row, column=3, value=lead.company)
        ws.cell(row=row, column=4, value=lead.email)
        ws.cell(row=row, column=5, value=lead.linkedin_url)
        ws.cell(row=row, column=6, value=lead.ai_score)
        ws.cell(row=row, column=7, value=lead.status)

    # auto-size columns
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{campaign.name}_leads.xlsx"'
    wb.save(response)
    return response
